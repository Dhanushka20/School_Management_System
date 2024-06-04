import tkinter as tk
from tkinter import messagebox, filedialog
from openpyxl import Workbook, load_workbook

class Person:
    def __init__(self, person_id, first_name, last_name):
        self.person_id = person_id
        self.first_name = first_name
        self.last_name = last_name

    def get_person_id(self):
        return self.person_id

    def get_first_name(self):
        return self.first_name

    def get_last_name(self):
        return self.last_name

    def set_first_name(self, first_name):
        self.first_name = first_name

    def set_last_name(self, last_name):
        self.last_name = last_name

    def display_info(self):
        return f"ID: {self.person_id}, Name: {self.first_name} {self.last_name}"

class Student(Person):
    def __init__(self, person_id, first_name, last_name, age, grade):
        super().__init__(person_id, first_name, last_name)
        self.age = age
        self.grade = grade

    def get_age(self):
        return self.age

    def get_grade(self):
        return self.grade

    def display_info(self):
        return super().display_info() + f", Age: {self.age}, Grade: {self.grade}"

class Teacher(Person):
    def __init__(self, person_id, first_name, last_name, subject):
        super().__init__(person_id, first_name, last_name)
        self.subject = subject

    def get_subject(self):
        return self.subject

    def display_info(self):
        return super().display_info() + f", Subject: {self.subject}"

class School:
    def __init__(self):
        self.students = []
        self.teachers = []
        self.workbook = Workbook()
        self.student_sheet = self.workbook.active
        self.student_sheet.title = "Students"
        self.student_sheet.append(["ID", "First Name", "Last Name", "Age", "Grade"])
        self.teacher_sheet = self.workbook.create_sheet("Teachers")
        self.teacher_sheet.append(["ID", "First Name", "Last Name", "Subject"])

    def add_student(self, student):
        self.students.append(student)
        self.student_sheet.append([student.person_id, student.first_name, student.last_name, student.age, student.grade])

    def add_teacher(self, teacher):
        self.teachers.append(teacher)
        self.teacher_sheet.append([teacher.person_id, teacher.first_name, teacher.last_name, teacher.subject])

    def display_all_students(self):
        return [student.display_info() for student in self.students]

    def display_all_teachers(self):
        return [teacher.display_info() for teacher in self.teachers]

    def find_student_by_id(self, student_id):
        for student in self.students:
            if student.get_person_id() == student_id:
                return student.display_info()
        return "Student not found."

    def find_teacher_by_id(self, teacher_id):
        for teacher in self.teachers:
            if teacher.get_person_id() == teacher_id:
                return teacher.display_info()
        return "Teacher not found."

    def save_to_excel(self):
        # Open a file dialog to let the user choose the save location
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                 filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if file_path:
            try:
                self.workbook.save(file_path)
                messagebox.showinfo("Success", f"Data saved to Excel file: {file_path}")
            except Exception as e:
                messagebox.showerror("Error", f"An error occurred while saving the file: {e}")

def add_student():
    try:
        student_id = int(entry_student_id.get())
        first_name = entry_student_fname.get()
        last_name = entry_student_lname.get()
        age = int(entry_student_age.get())
        grade = entry_student_grade.get()
        school.add_student(Student(student_id, first_name, last_name, age, grade))
        school.save_to_excel()
    except ValueError:
        messagebox.showerror("Error", "Please enter valid data for all fields.")

def add_teacher():
    try:
        teacher_id = int(entry_teacher_id.get())
        first_name = entry_teacher_fname.get()
        last_name = entry_teacher_lname.get()
        subject = entry_teacher_subject.get()
        school.add_teacher(Teacher(teacher_id, first_name, last_name, subject))
        school.save_to_excel()
    except ValueError:
        messagebox.showerror("Error", "Please enter valid data for all fields.")

def display_students():
    messagebox.showinfo("Students", "\n".join(school.display_all_students()))

def display_teachers():
    messagebox.showinfo("Teachers", "\n".join(school.display_all_teachers()))

def find_student():
    try:
        student_id = int(entry_find_student_id.get())
        messagebox.showinfo("Student Info", school.find_student_by_id(student_id))
    except ValueError:
        messagebox.showerror("Error", "Please enter a valid student ID.")

def find_teacher():
    try:
        teacher_id = int(entry_find_teacher_id.get())
        messagebox.showinfo("Teacher Info", school.find_teacher_by_id(teacher_id))
    except ValueError:
        messagebox.showerror("Error", "Please enter a valid teacher ID.")

def quit_program():
    root.destroy()

school = School()

root = tk.Tk()
root.title("School Management System")

# Bold font
bold_font = ('Helvetica', 10, 'bold')

# Add Student Section
tk.Label(root, text="Add Student", font=bold_font).grid(row=0, column=0, columnspan=2, pady=5)
tk.Label(root, text="ID:", font=bold_font).grid(row=1, column=0, pady=5)
entry_student_id = tk.Entry(root, font=bold_font)
entry_student_id.grid(row=1, column=1, pady=5)
tk.Label(root, text="First Name:", font=bold_font).grid(row=2, column=0, pady=5)
entry_student_fname = tk.Entry(root, font=bold_font)
entry_student_fname.grid(row=2, column=1, pady=5)
tk.Label(root, text="Last Name:", font=bold_font).grid(row=3, column=0, pady=5)
entry_student_lname = tk.Entry(root, font=bold_font)
entry_student_lname.grid(row=3, column=1, pady=5)
tk.Label(root, text="Age:", font=bold_font).grid(row=4, column=0, pady=5)
entry_student_age = tk.Entry(root, font=bold_font)
entry_student_age.grid(row=4, column=1, pady=5)
tk.Label(root, text="Grade:", font=bold_font).grid(row=5, column=0, pady=5)
entry_student_grade = tk.Entry(root, font=bold_font)
entry_student_grade.grid(row=5, column=1, pady=5)
tk.Button(root, text="Add Student", command=add_student, bg='blue', fg='white', font=bold_font).grid(row=6, column=0, columnspan=2, pady=5)

# Add Teacher Section
tk.Label(root, text="Add Teacher", font=bold_font).grid(row=7, column=0, columnspan=2, pady=5)
tk.Label(root, text="ID:", font=bold_font).grid(row=8, column=0, pady=5)
entry_teacher_id = tk.Entry(root, font=bold_font)
entry_teacher_id.grid(row=8, column=1, pady=5)
tk.Label(root, text="First Name:", font=bold_font).grid(row=9, column=0, pady=5)
entry_teacher_fname = tk.Entry(root, font=bold_font)
entry_teacher_fname.grid(row=9, column=1, pady=5)
tk.Label(root, text="Last Name:", font=bold_font).grid(row=10, column=0, pady=5)
entry_teacher_lname = tk.Entry(root, font=bold_font)
entry_teacher_lname.grid(row=10, column=1, pady=5)
tk.Label(root, text="Subject:", font=bold_font).grid(row=11, column=0, pady=5)
entry_teacher_subject = tk.Entry(root, font=bold_font)
entry_teacher_subject.grid(row=11, column=1, pady=5)
tk.Button(root, text="Add Teacher", command=add_teacher, bg='green', fg='white', font=bold_font).grid(row=12, column=0, columnspan=2, pady=5)

# Display Students Section
tk.Button(root, text="Display Students", command=display_students, bg='orange', fg='white', font=bold_font).grid(row=13, column=0, columnspan=2, pady=5)

# Display Teachers Section
tk.Button(root, text="Display Teachers", command=display_teachers, bg='orange', fg='white', font=bold_font).grid(row=14, column=0, columnspan=2, pady=5)

# Find Student Section
tk.Label(root, text="Find Student by ID", font=bold_font).grid(row=15, column=0, columnspan=2, pady=5)
tk.Label(root, text="Student ID:", font=bold_font).grid(row=16, column=0, pady=5)
entry_find_student_id = tk.Entry(root, font=bold_font)
entry_find_student_id.grid(row=16, column=1, pady=5)
tk.Button(root, text="Find Student", command=find_student, bg='purple', fg='white', font=bold_font).grid(row=17, column=0, columnspan=2, pady=5)

# Find Teacher Section
tk.Label(root, text="Find Teacher by ID", font=bold_font).grid(row=18, column=0, columnspan=2, pady=5)
tk.Label(root, text="Teacher ID:", font=bold_font).grid(row=19,column=0, pady=5)
entry_find_teacher_id = tk.Entry(root, font=bold_font)
entry_find_teacher_id.grid(row=19, column=1, pady=5)
tk.Button(root, text="Find Teacher", command=find_teacher, bg='purple', fg='white', font=bold_font).grid(row=20, column=0, columnspan=2, pady=5)

# Quit Program Section
tk.Button(root, text="Quit", command=quit_program, bg='red', fg='white', font=bold_font).grid(row=21, column=0, columnspan=2, pady=5)

root.mainloop()
